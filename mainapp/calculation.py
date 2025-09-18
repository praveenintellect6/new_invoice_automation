
from .models import *
import pandas as pd
import operator
import re 
import numpy as np
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import shutil
from .purchase_report import PurchaseReportClass
from datetime import datetime

class ReportCalculation:
    #*kwargs(df=df)
    numeric_column=['TRADE PRICE','TOTAL COUNT','PURCHASED COUNT','TOTAL PRICE','ACTUAL PRICE','SELLING PRICE(Exc.GST)','GST','SELLING PRICE(Inc.GST)']
    def __init__(self, file=None,filename=None,df=None,excel_date=None,supp=None):
        self.supp=supp
        self.operator=["+", "-", "*", "/"]
        #'0','1','2','3','4','5','6','7','8','9'
        # self.supp=NewSupplier.objects.get(id=24)
        self.cases=Cases.objects.filter(supplier=self.supp).values('min','max','profit')
        self.gst_profit_case=Cases.objects.filter(supplier=self.supp).first()
        self.gst=self.gst_profit_case.gst

        self.gst_mapp=self.gst_profit_case.gst_mapp
        self.profit_mapp=self.gst_profit_case.profit_mapp
        
        self.supplier_name=self.supp.supplier_name
        self.mapped_col = self.supp.supplier_mapp_col
        self.col_list = self.supp.supplier_col
        self.equation=[]
        self.direct=[]
        load_dotenv()
        self.template_path=os.getenv("workbook_path")
        self.equation_token_mapp={}
        self.extractExpression()
        #---------------------------
        #self.read_excel()
        if df is not None:
            self.df=df
        if file is not None:
            self.filepath=file

        if filename is not None:
            self.filename=filename

        if excel_date is not None:
            self.excel_date=excel_date

    def get_equation(self):
        return self.equation
    
    def get_direct(self):
        return self.direct
    
    def extractExpression(self):
        self.sorted_data = {key: self.mapped_col.get(key,"") for key in default_mapping().keys()}
        for key, value in self.sorted_data.items():
            if any(op in value for op in self.operator):
                self.equation.append(key)
            else:
                self.direct.append(key)

    
    def read_excel(self,):
        df = pd.read_excel("196_Invoice_4321420874.xlsx")
        self.df=df
    
    def exportToExcel(self):
        self.sample=pd.DataFrame()
        for key in default_mapping().keys():
            self.sample[key] = self.combine_report[key]

        new_date = datetime.strptime(self.excel_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        self.sample['DATE'] = new_date
        self.sample['PROFIT%'] = self.combine_report['PROFIT_per']
        for i in self.sample.columns:
            if i not in self.numeric_column:
                self.sample[i]=self.sample[i].astype(str)
            else:
                self.sample[i]=self.sample[i].astype(float)

        filepathname = os.path.join(self.filepath, f"{self.excel_date}_PurchaseReport.xlsx")

        if os.path.exists(filepathname):
            self.wb = load_workbook(filepathname)
        else:
            self.wb = load_workbook(self.template_path, data_only=True)
            
        ws = self.wb.active
        headers = [cell.value for cell in ws[4]] 
        start_row = ws.max_row + 1
        for r_idx, row in enumerate(self.sample.to_dict('records'), start=start_row):
            for c_idx, header in enumerate(headers, start=1):
                if header == "S.NO":
                    ws.cell(row=r_idx, column=c_idx, value=r_idx - 4)
                elif header in row:
                    ws.cell(row=r_idx, column=c_idx, value=row[header])

        print('Purchase Report saving............!')
        self.wb.save(filepathname)
        self.wb.close()
        return self.sample

    def getProcessedReport(self):
        return self.sample

    def findProfit(self):
        def findpro(value):
            value=float(value)
            for i in self.cases:
                if int(i['min']) < value <= int(i['max']):
                    return round(float(i['profit'])/100 * value,2)
            return None
        
        def findper(value):
            value=float(value)
            for i in self.cases:
                if int(i['min']) < value <= int(i['max']):
                    return i['profit']+' %'
            return None
        
        self.combine_report["PROFIT%"] = self.combine_report[self.profit_mapp].apply(findpro)
        self.combine_report["PROFIT_per"] = self.combine_report[self.profit_mapp].apply(findper)

    def findGst(self):
        def findgst(value):
            value=float(value)
            gst=float(self.gst)/100
            return round((value * gst),2)
        self.combine_report["GST"] = self.combine_report[self.gst_mapp].apply(findgst)

    def MappToPurchaseReport(self):
        self.report=pd.DataFrame(columns=list(self.mapped_col.keys()))
        mergelst=list(self.mapped_col.keys()) + (list(self.df.columns))
        self.combine_report=pd.DataFrame(columns=mergelst)

        for i in list(self.df.columns):
            self.combine_report[i]=self.df[i]
        
        if self.gst_mapp in self.combine_report.columns:
            self.findProfit()

        if self.profit_mapp in self.combine_report.columns:
            self.findGst()

        for key,value in self.sorted_data.items():
            if key in self.direct:
                    if value in self.df.columns:
                        self.report[key] = self.df[value]
                        self.combine_report[key]=self.df[value]

            if key in self.equation:
                sorted_cols = sorted(list(self.combine_report.columns), key=len, reverse=True)
                escaped_cols = [re.escape(col) for col in sorted_cols]
                pattern = r'(' + '|'.join(escaped_cols) + r'|\+|\-|\*|\/|\(|\)|%|\d+(?:\.\d+)?)'
                tokens = [t.strip() for t in re.findall(pattern, value) if t.strip()]
                tokens = ' '.join([f"pd.to_numeric(self.combine_report['{i}'], errors='coerce')" if i in list(self.combine_report.columns) else i for i in tokens])
                self.equation_token_mapp[key]=tokens
                self.combine_report[key] =  eval(self.equation_token_mapp[key]).round(2)

            if key == self.profit_mapp:
                    self.findProfit()

            if key == self.gst_mapp:
                    self.findGst()

        print("-----------------calculation end---------------------------------")
    

