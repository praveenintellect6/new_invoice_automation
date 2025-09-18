from datetime import datetime
from django.conf import settings
import os 
import shutil
import time
from dotenv import load_dotenv
from openpyxl import load_workbook
import pandas as pd
from .models import *


class PurchaseReportClass:
    def __init__(self):
        pass

    def createFolderByDate(self,date_str):
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        folder_path = os.path.join(settings.MEDIA_ROOT,str(dt.year),dt.strftime("%b"),date_str)
        os.makedirs(folder_path, exist_ok=True)
        return folder_path
    
    def createFolderFromMedia(self, date_str):
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        folder_path = os.path.join('media',str(dt.year),dt.strftime("%b"),date_str)
        return folder_path
    
    def copy_folder_contents(source_folder,retries=5, delay=0.5):
        if not os.path.exists(source_folder):
            print(f"Source folder does not exist: {source_folder}")
            return
        relative_path = os.path.relpath(source_folder)
        load_dotenv()
        remote_location=os.getenv("remote_location")
        #destination_folder=rf"\\system26\D\Jijo\newmedia\{relative_path}"
        destination_folder=os.path.join(remote_location,relative_path)
        os.makedirs(destination_folder, exist_ok=True)

        for root, dirs, files in os.walk(source_folder):
            relative_path = os.path.relpath(root, source_folder)
            dest_root = os.path.join(destination_folder, relative_path)
            os.makedirs(dest_root, exist_ok=True)

            for file in files:
                src_file = os.path.join(root, file)
                dst_file = os.path.join(dest_root, file)

                for attempt in range(retries):
                    try:
                        shutil.copy2(src_file, dst_file)  
                        break
                    except PermissionError:
                        print(f" Locked, retrying ({attempt+1}/{retries}): {src_file}")
                        time.sleep(delay)
                else:
                    print(f" Skipped locked file: {src_file}")
    
    def monthlyPurchaseReport(self,file_list=None,year=None,month=None):
        all_dfs = []
        for f in file_list:
            df = pd.read_excel(f,header=3)
            all_dfs.append(df)
        self.final_df = pd.concat(all_dfs, ignore_index=True)

        self.final_df =  self.final_df.drop(columns=['S.NO'])
        self.final_df.insert(0, 'S.NO', range(1, len( self.final_df) + 1))

        self.template_path=os.getenv("workbook_path")
        self.wb = load_workbook(self.template_path, data_only=True)
        ws = self.wb.active
        headers = [cell.value for cell in ws[4]]
        start_row = ws.max_row + 1

        for r_idx, row in enumerate(self.final_df.to_dict('records'), start=start_row):
            for c_idx, header in enumerate(headers, start=1):
                if header == "S.NO":
                    ws.cell(row=r_idx, column=c_idx, value=r_idx - 4)
                elif header in row:
                    ws.cell(row=r_idx, column=c_idx, value=row[header])

        folder_path= os.path.join('media',year,month)
        file_path = os.path.join(folder_path, f"PurchaseReport{month}_{year}.xlsx")
        self.wb.save(file_path)
        self.wb.close()
    
    def autoSelectingSupplier(self,df=None):
        self.nn=NewSupplier.objects.all()
        self.supplier_name=None
        self.df_col = df.columns.tolist()
        for i in self.nn:
            equal = set(x.lower() for x in self.df_col) == set(x.lower() for x in i.supplier_col)
            if equal:
                self.supplier_name=i.supplier_name
            else:
                continue
        return self.supplier_name
    
    def autoSelectingSupplierByRow(self,df=None):
        self.df=df
        self.first_row = df.iloc[0,:].tolist()
        self.first_row = [str(x) if pd.isna(x) else x for x in self.first_row]
        self.supplier_name_list = list(NewSupplier.objects.values_list('supplier_name', flat=True))
        self.suppliername = next((item for item in self.first_row if item in self.supplier_name_list), None)
        self.exist = self.suppliername is not None
        self.col = df.columns.tolist()
        self.report_col=default_mapping().keys()
        common_items = list(set(self.col) & set(self.report_col))
        for i in common_items:
            self.df = self.df.drop(columns=[i])
        return self.suppliername,self.df
    
    

        

        


        


        


          

            

    
 


